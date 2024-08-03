import random  # 导入随机模块，用于随机选择语言
import time  # 导入时间模块，用于控制延迟
import configparser  # 导入配置文件解析模块，用于读取配置文件
from datetime import datetime  # 导入日期时间模块，用于日志文件命名
from deep_translator import GoogleTranslator  # 导入Google翻译模块
import logging  # 导入日志模块，用于记录日志
import colorlog  # 导入彩色日志模块，用于在控制台显示彩色日志
import os  # 导入操作系统模块，用于文件操作
import pandas as pd  # 导入Pandas模块，用于数据处理和Excel文件导出
from openpyxl import load_workbook  # 导入openpyxl库，用于Excel文件处理
from openpyxl.utils import get_column_letter  # 导入openpyxl工具，用于列宽调整
from openpyxl.styles import Alignment, Font  # 导入openpyxl样式，用于单元格格式化

# 设置日志记录器
logger = logging.getLogger('my_logger')  # 创建日志记录器对象
logger.setLevel(logging.DEBUG)  # 设置日志记录器级别为DEBUG

# 设置日志格式
formatter = colorlog.ColoredFormatter(
    "%(log_color)s [%(asctime)s] -%(levelname)s- %(message)s",
    datefmt=None,
    reset=True,
    log_colors={
        'DEBUG': 'cyan',  # DEBUG级别日志显示为青色
        'INFO': 'green',  # INFO级别日志显示为绿色
        'WARNING': 'yellow',  # WARNING级别日志显示为黄色
        'ERROR': 'red',  # ERROR级别日志显示为红色
        'CRITICAL': 'bold_red',  # CRITICAL级别日志显示为粗红色
    }
)

# 创建文件日志格式
fileformatter = colorlog.ColoredFormatter(
    "%(log_color)s [%(asctime)s] (%(threadName)s) -%(levelname)s- %(message)s",
)

# 创建控制台处理器并设置级别和格式
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.DEBUG)
console_handler.setFormatter(formatter)

# 获取当前时间并格式化，用于日志文件命名
current_time = datetime.now()
formatted_time = current_time.strftime("%Y-%m-%d,%H-%M-%S")

# 检查日志目录是否存在，不存在则创建
if not os.path.exists("logs"):
    try:
        os.makedirs("logs")  # 创建目录
    except Exception as e:
        logger.critical("无法创建目录 logs")
        raise SystemExit
    logger.info("已创建目录 logs")
else:
    logger.info("logs 目录已存在")

# 创建文件处理器并设置级别和格式
file_handler = logging.FileHandler("logs/" + str(formatted_time) + ".log", encoding="utf-8", mode="w+")
file_handler.setLevel(logging.DEBUG)
file_handler.setFormatter(fileformatter)

# 将处理器添加到日志记录器
logger.addHandler(console_handler)
logger.addHandler(file_handler)

# 创建或读取配置文件
config = configparser.ConfigParser()

if not os.path.exists("config.ini"):
    # 如果配置文件不存在，则创建一个默认配置文件
    config_file_content = '''\
[options]

# 待生草的文件
file_src = src.txt

# 生草结果输出文件
file_out = translation_results.xlsx 

# 编码
encoding = utf-8

# 目标语言（详见源代码内的language_names映射）
target_lang = zh-CN

# 生草次数
frequency = 20
'''

    with open('config.ini', 'w') as configfile:
        configfile.write(config_file_content)
    logger.info("配置文件缺失，已自动生成 config.ini")
    logger.warning(f"请配置config.ini后再次运行程序。")
    raise SystemExit
else:
    config.read("config.ini")
    logger.info("配置文件读取成功")

enc = config["options"]["encoding"]  # 从配置文件中读取编码设置

# 检查源文件是否存在，如果不存在则创建一个空白文件并提示用户填写内容
file_src = config["options"]["file_src"]
if not os.path.exists(file_src):
    with open(file_src, 'w', encoding=enc) as f:
        f.write("")  # 创建一个空白文件
    logger.warning(f"{file_src} 文件不存在，已创建空白文件，请填写需要翻译的内容后再运行程序。")
    raise SystemExit

# 创建GoogleTranslator实例
translator = GoogleTranslator()

# 获取支持的语言列表
languages = translator.get_supported_languages(as_dict=True)  # 获取支持的语言列表，返回字典格式
languages_list = list(languages.values())  # 提取语言代码列表

# 语言代码到中文名称的映射
language_names = {
    'af': '南非荷兰语', 'sq': '阿尔巴尼亚语', 'am': '阿姆哈拉语', 'ar': '阿拉伯语', 'hy': '亚美尼亚语', 'as': '阿萨姆语',
    'ay': '艾马拉语', 'az': '阿塞拜疆语', 'bm': '班巴拉语', 'eu': '巴斯克语', 'be': '白俄罗斯语', 'bn': '孟加拉语',
    'bho': '博杰普尔语', 'bs': '波斯尼亚语', 'bg': '保加利亚语', 'ca': '加泰罗尼亚语', 'ceb': '宿务语', 'ny': '尼扬贾语',
    'zh-CN': '中文（简体）', 'zh-TW': '中文（繁体）', 'co': '科西嘉语', 'hr': '克罗地亚语', 'cs': '捷克语', 'da': '丹麦语',
    'dv': '迪维希语', 'doi': '多格来语', 'nl': '荷兰语', 'en': '英语', 'eo': '世界语', 'et': '爱沙尼亚语', 'ee': '埃维语',
    'tl': '菲律宾语', 'fi': '芬兰语', 'fr': '法语', 'fy': '弗里西语', 'gl': '加利西亚语', 'ka': '格鲁吉亚语', 'de': '德语',
    'el': '希腊语', 'gn': '瓜拉尼语', 'gu': '古吉拉特语', 'ht': '海地克里奥尔语', 'ha': '豪萨语', 'haw': '夏威夷语',
    'iw': '希伯来语', 'hi': '印地语', 'hmn': '苗语', 'hu': '匈牙利语', 'is': '冰岛语', 'ig': '伊博语', 'ilo': '伊洛卡诺语',
    'id': '印尼语', 'ga': '爱尔兰语', 'it': '意大利语', 'ja': '日语', 'jw': '爪哇语', 'kn': '卡纳达语', 'kk': '哈萨克语',
    'km': '高棉语', 'rw': '基尼阿万达语', 'gom': '贡根语', 'ko': '韩语', 'kri': '克里奥语', 'ku': '库尔德语（库尔曼吉）',
    'ckb': '库尔德语（索拉尼）', 'ky': '吉尔吉斯语', 'lo': '老挝语', 'la': '拉丁语', 'lv': '拉脱维亚语', 'ln': '林加拉语',
    'lt': '立陶宛语', 'lg': '卢干达语', 'lb': '卢森堡语', 'mk': '马其顿语', 'mai': '迈蒂利语', 'mg': '马尔加什语',
    'ms': '马来语', 'ml': '马拉雅拉姆语', 'mt': '马耳他语', 'mi': '毛利语', 'mr': '马拉地语', 'mni-Mtei': '梅泰语',
    'lus': '米佐语', 'mn': '蒙古语', 'my': '缅甸语', 'ne': '尼泊尔语', 'no': '挪威语', 'or': '奥利亚语', 'om': '奥罗莫语',
    'ps': '普什图语', 'fa': '波斯语', 'pl': '波兰语', 'pt': '葡萄牙语', 'pa': '旁遮普语', 'qu': '克丘亚语', 'ro': '罗马尼亚语',
    'ru': '俄语', 'sm': '萨摩亚语', 'sa': '梵语', 'gd': '苏格兰盖尔语', 'nso': '北索托语', 'sr': '塞尔维亚语', 'st': '塞索托语',
    'sn': '绍纳语', 'sd': '信德语', 'si': '僧伽罗语', 'sk': '斯洛伐克语', 'sl': '斯洛文尼亚语', 'so': '索马里语',
    'es': '西班牙语', 'su': '巽他语', 'sw': '斯瓦希里语', 'sv': '瑞典语', 'tg': '塔吉克语', 'ta': '泰米尔语', 'tt': '鞑靼语',
    'te': '泰卢固语', 'th': '泰语', 'ti': '提格里尼亚语', 'ts': '聪加语', 'tr': '土耳其语', 'tk': '土库曼语', 'ak': '阿肯语',
    'uk': '乌克兰语', 'ur': '乌尔都语', 'ug': '维吾尔语', 'uz': '乌兹别克语', 'vi': '越南语', 'cy': '威尔士语', 'xh': '科萨语',
    'yi': '意第绪语', 'yo': '约鲁巴语', 'zu': '祖鲁语'
}

# 定义一个数组打乱函数，用于随机选择目标语言
def shuffle_array(arr):
    n = len(arr)  # 获取数组长度
    for i in range(n - 1, 0, -1):  # 从数组末尾开始迭代
        time.sleep(0.05)  # 暂停一小段时间，模拟处理延迟
        j = random.randint(0, i)  # 生成一个随机索引
        arr[i], arr[j] = arr[j], arr[i]  # 交换元素
    return arr

# 尝试打开源文件
try:
    file = open(file_src, "r", encoding=enc)  # 以只读模式打开源文件
    logger.info("源文件读取成功")
except Exception as e:
    logger.error(e)
    logger.critical("源文件读取出错，请检查配置文件!")
    raise SystemExit

content = file.read().strip()  # 读取文件内容并去除首尾空格
file.close()  # 关闭文件

translation_data = []  # 创建一个列表，用于存储翻译数据

# 定义翻译函数
def translator_(freq, content):
    text = content  # 获取需要翻译的文本
    for i in range(freq):  # 循环翻译指定次数
        targetlang = shuffle_array(languages_list)[0]  # 随机选择目标语言
        while True:
            try:
                text = GoogleTranslator(source="auto", target=targetlang).translate(text)  # 翻译文本
                break
            except Exception as e:
                logger.warning(e)
                continue
        logger.info(f"翻译第 {i + 1} 次:" + text)
        text_zh = GoogleTranslator(source="auto", target="zh-CN").translate(text)  # 将翻译后的文本再翻译成简体中文
        translation_data.append([i + 1, targetlang, language_names.get(targetlang, '未知语言'), text, text_zh])  # 将翻译结果存储到列表中
    text = GoogleTranslator(source="auto", target=config["options"]["target_lang"]).translate(text)  # 最终将文本翻译为目标语言
    translation_data.append([freq, config["options"]["target_lang"], language_names.get(config["options"]["target_lang"], '未知语言'), text, text])  # 将最终结果存储到列表中
    logger.info(f"翻译完成")

# 执行翻译
f = int(config["options"]["frequency"])  # 从配置文件中读取翻译次数
translator_(f, content)  # 调用翻译函数

# 创建DataFrame并导出为Excel文件
df = pd.DataFrame(translation_data, columns=["生草次数", "目标语言", "目标语言（中文全称）", "结果", "结果翻译成中文（简体）"])
df.to_excel(config["options"]["file_out"], index=False)  # 将翻译结果保存为Excel文件
logger.info(f"翻译结果已保存为 {config['options']['file_out']}")

# 加载保存的Excel文件，进行格式化操作
wb = load_workbook(config["options"]["file_out"])
ws = wb.active

# 设置首行居中
for cell in ws[1]:
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)

# 自动换行和自动适应列宽
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # 获取列字母
    for cell in col:
        cell.alignment = Alignment(wrap_text=True)  # 设置单元格自动换行
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))  # 获取列中最长单元格的长度
    adjusted_width = (max_length + 2)  # 设置适应列宽
    ws.column_dimensions[column].width = adjusted_width

# 保存格式化后的Excel文件
wb.save(config["options"]["file_out"])
logger.info("Excel文件格式化完成!")
